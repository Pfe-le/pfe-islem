import sqlite3
import pandas as pd
import openpyxl
import os

# --- chemins des fichiers ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "gestion_production.db")
KANBAN_PATH = os.path.join(BASE_DIR, "Classeur Kanban VKF CW 12.xlsm")
PDB_PATH = os.path.join(BASE_DIR, "LAS_PDB .xlsm")
# --------------------------------
def init_db():
    #1.connexion
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()


    # 2. Reset Tables
    cursor.execute("DROP TABLE IF EXISTS Produits")
    cursor.execute("DROP TABLE IF EXISTS Stock")
    cursor.execute("DROP TABLE IF EXISTS Demandes")
    cursor.execute("DROP TABLE IF EXISTS Pannes") # Supprimer pour mettre à jour la structure

    # 3. Création des tables
    cursor.execute("""
    CREATE TABLE Pannes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        operateur_id TEXT,
        machine_id TEXT,
        cause TEXT NOT NULL,
        debut_panne TEXT NOT NULL,
        fin_panne TEXT,
        statut TEXT DEFAULT '🔴 Ouvert'
    )
    """)

    cursor.execute("""
    CREATE TABLE Produits (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        reference TEXT UNIQUE,
        famille TEXT,
        module TEXT,
        pression REAL,
        temps REAL,
        amplitude REAL
    )
    """)

    cursor.execute("CREATE TABLE Stock (reference TEXT PRIMARY KEY, famille TEXT, quantite INTEGER DEFAULT 0)")
    
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS Demandes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        reference TEXT,
        quantite INTEGER,
        date_besoin TEXT,
        shift TEXT,
        statut TEXT DEFAULT '🟠En attente',
        urgence TEXT,
        heure_demande TEXT,
        debut_production TEXT,
        fin_production TEXT,
        operateur_id TEXT
    )
    """)

    # 4. Import Dispatching
    try:
        df_k = pd.read_excel(KANBAN_PATH, sheet_name="DISPATCHING REF")
        for i, row in df_k.iterrows():
            try:
                ref = str(row.iloc[2]).strip()
                if ref.lower() not in ['nan', '', 'none', 'ref cab']:
                    cursor.execute("INSERT OR IGNORE INTO Produits (reference, famille, module) VALUES (?, ?, ?)", 
                                 (ref, str(row.iloc[1]).strip(), str(row.iloc[0]).strip()))
                    cursor.execute("INSERT OR IGNORE INTO Stock (reference, famille, quantite) VALUES (?, ?, 0)", 
                                 (ref, str(row.iloc[1]).strip()))
            except:
                continue
        print("✅ Dispatching OK")
    except Exception as e:
        print(f"❌ Erreur Dispatching: {e}")

    # 5. Import BESOIN
    # --- 4. Import BESOIN ---
    try:
        wb = openpyxl.load_workbook(KANBAN_PATH, data_only=True)
        if "BESOIN" in wb.sheetnames:
            sheet = wb["BESOIN"]
            count_log = 0

        # Parcours des lignes 2 à 500 (ajuste si plus de lignes)
            for row_idx in range(2, 501):
                for col_idx in range(1, 4):  # A=Fiat, B=Fors, C=Sigip
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None:
                        val_ref = str(cell_value).strip()

                    # Ignorer les entêtes / valeurs inutiles
                        if val_ref.lower() not in ['nan','none','','fiat pn','fors pn','sigip pn','ref cab','ref','pn']:
                        
                        # Vérifie si la référence existe déjà
                            cursor.execute("SELECT reference FROM Produits WHERE reference=?", (val_ref,))
                            exists = cursor.fetchone()

                            if not exists:
                            # Ajouter seulement la référence, module/famille restent par défaut
                                cursor.execute("""
                                    INSERT INTO Produits (reference, famille, module)
                                    VALUES (?, 'Reference_Cable', 'LOGISTIQUE')
                                """, (val_ref,))
                            
                                cursor.execute("""
                                    INSERT INTO Stock (reference, famille, quantite)
                                    VALUES (?, 'Reference_Cable', 0)
                                """, (val_ref,))
                            
                                count_log += 1

            print(f"✅ Logistique OK : {count_log} références lues dans Fiat, Fors et Sigip.")
        else:
            print("❌ Onglet 'BESOIN' introuvable.")
    except Exception as e:
        print(f"❌ Erreur Logistique : {e}")

    # 6. Import PDB
    try:
        df_pdb = pd.read_excel(PDB_PATH, sheet_name=0)
        for _, row in df_pdb.iterrows():
            ref_p = str(row.iloc[1]).strip()
            if ref_p not in ['nan', '']:
                cursor.execute("""
                    UPDATE Produits 
                    SET pression=?, temps=?, amplitude=? 
                    WHERE reference=?
                """, (
                    pd.to_numeric(row.iloc[2], errors='coerce'), 
                    pd.to_numeric(row.iloc[5], errors='coerce'), 
                    pd.to_numeric(row.iloc[6], errors='coerce'), 
                    ref_p
                ))
        print("✅ PDB OK")
    except Exception as e:
        print(f"❌ Erreur PDB: {e}")

    # 7. Commit & fermeture
    conn.commit()
    conn.close()
    print("✅ Database terminée !")

if __name__ == "__main__":
    init_db()